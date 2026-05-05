"""
DataParser Pro - B2B Order Translator
Utilitário para conversão de relatórios desestruturados de pedidos em tabelas relacionais.
"""

import os
import threading
import time
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox

import customtkinter as ctk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class TradutorPremiumApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("DataParser Pro")
        self.geometry("500x380")
        self.resizable(False, False)
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.card = ctk.CTkFrame(self, corner_radius=15, fg_color=("gray90", "gray13"))
        self.card.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

        self.label_titulo = ctk.CTkLabel(
            self.card, text="Conversor de Pedidos", 
            font=ctk.CTkFont(family="Segoe UI", size=22, weight="bold")
        )
        self.label_titulo.pack(pady=(30, 5))

        self.label_sub = ctk.CTkLabel(
            self.card, text="Transforme relatórios em dados estruturados.", 
            font=ctk.CTkFont(family="Segoe UI", size=12), text_color="gray"
        )
        self.label_sub.pack(pady=(0, 25))

        self.btn_processar = ctk.CTkButton(
            self.card, text="Selecionar Arquivo", 
            command=self.iniciar_processo, 
            height=45, width=220, corner_radius=8,
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            fg_color="#6A5ACD", hover_color="#483D8B"
        )
        self.btn_processar.pack(pady=15)

        self.progress_bar = ctk.CTkProgressBar(self.card, width=220, height=8, corner_radius=4, progress_color="#6A5ACD")
        self.progress_bar.set(0)
        self.progress_bar.pack(pady=(10, 0))
        self.progress_bar.pack_forget()

        self.label_status = ctk.CTkLabel(
            self.card, text="Aguardando importação...", 
            font=ctk.CTkFont(family="Segoe UI", size=11), text_color="gray"
        )
        self.label_status.pack(pady=(15, 10))

    def iniciar_processo(self):
        caminho_entrada = filedialog.askopenfilename(title="Selecionar Pedidos", filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")])
        if not caminho_entrada: 
            return

        self.btn_processar.configure(state="disabled", text="Processando...")
        self.progress_bar.pack(pady=(10, 0))
        self.progress_bar.start()
        self.label_status.configure(text="Lendo dados brutos...", text_color="orange")
        
        threading.Thread(target=self.fluxo_principal, args=(caminho_entrada,)).start()

    def fluxo_principal(self, caminho_entrada):
        try:
            time.sleep(0.5) 
            df_final = self.processar_dados(caminho_entrada)
            
            self.progress_bar.stop()
            self.progress_bar.set(1)
            self.label_status.configure(text="Validação concluída.", text_color="green")

            total_pedidos = df_final['ID_PEDIDO'].nunique()
            total_clientes = df_final['CNPJ'].nunique()

            relatorio = (f"Resumo do Arquivo:\n\n"
                         f"• Pedidos Únicos: {total_pedidos}\n"
                         f"• Clientes Distintos: {total_clientes}\n"
                         f"• Total de Itens: {len(df_final)}\n\n"
                         f"Deseja salvar a versão estruturada?")

            self.after(0, self.confirmar_salvamento, df_final, caminho_entrada, relatorio)

        except Exception as e:
            self.progress_bar.stop()
            self.progress_bar.pack_forget()
            self.after(0, lambda: messagebox.showerror("Erro", f"Falha ao processar:\n{str(e)}"))
            self.reset_ui()

    def confirmar_salvamento(self, df_final, caminho_entrada, relatorio):
        confirmar = messagebox.askyesno("Validar Pedidos", relatorio)
        if confirmar:
            self.label_status.configure(text="Salvando e formatando Excel...", text_color="orange")
            self.progress_bar.start()
            self.salvar_e_formatar(df_final, caminho_entrada)
        else:
            self.reset_ui()
            self.label_status.configure(text="Operação cancelada.", text_color="gray")

    def reset_ui(self):
        self.btn_processar.configure(state="normal", text="Selecionar Arquivo")
        self.progress_bar.stop()
        self.progress_bar.pack_forget()

    def processar_dados(self, caminho):
        if caminho.endswith('.csv'): 
            df_orig = pd.read_csv(caminho, header=None, dtype=str).fillna('')
        else: 
            df_orig = pd.read_excel(caminho, header=None, dtype=str).fillna('')

        dados = []
        pedido_id, cnpj, ie, bairro, cidade, uf, cep = [""] * 7
        lendo_itens = False

        for _, row in df_orig.iterrows():
            linha = [str(x).strip() for x in list(row)]
            c1 = linha[0]

            if "Pedido:" in c1:
                pedido_id = c1.replace("Pedido:", "").strip()
                lendo_itens = False
            elif "CNPJ:" in c1:
                texto_cnpj = " ".join([x for x in linha if x != ''])
                partes = texto_cnpj.split("IE:")
                cnpj = partes[0].replace("CNPJ:", "").strip()
                ie = partes[1].strip() if len(partes) > 1 else ""
                lendo_itens = False
            elif "Bairro:" in c1:
                pts = [x for x in linha if x != '']
                bairro, cidade, uf, cep = (pts + ["", "", "", ""])[:4]
                bairro = bairro.replace("Bairro:", "").strip()
                cidade = cidade.replace("Cidade:", "").strip()
                uf = uf.replace("UF:", "").strip()
                cep = cep.replace("Cep:", "").strip()
            elif "Código" in c1: 
                lendo_itens = True
            elif lendo_itens:
                if not c1 or "Procedimentos" in c1:
                    lendo_itens = False
                    continue
                itens = [x for x in linha if x != '']
                if len(itens) >= 3:
                    itens = (itens + ['0', '0', '0', '0', '0'])[:5]
                    dados.append([pedido_id, cnpj, ie, bairro, cidade, uf, cep] + itens)

        cols = ['ID_PEDIDO', 'CNPJ', 'IE', 'BAIRRO', 'CIDADE', 'UF', 'CEP', 'SKU_PEÇA', 'QUANTIDADE', 'PREÇO_UNIT', 'VLR_IPI', 'VLR_ST']
        return pd.DataFrame(dados, columns=cols)

    def salvar_e_formatar(self, df, caminho_orig):
        pasta = Path.home() / "Documents" / "Arquivos Traduzidos"
        pasta.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime('%d-%m-%Y_%H-%M')
        nome_base = os.path.splitext(os.path.basename(caminho_orig))[0]
        caminho_final = pasta / f"TRADUZIDO_{nome_base}_{timestamp}.xlsx"
        
        df.to_excel(caminho_final, index=False)

        wb = load_workbook(caminho_final)
        ws = wb.active
        for cell in ws[1]: 
            cell.font = Font(bold=True)
        ws.auto_filter.ref = ws.dimensions
        wb.save(caminho_final)

        self.reset_ui()
        self.label_status.configure(text="Arquivo salvo com sucesso!", text_color="green")
        os.startfile(pasta)

if __name__ == "__main__":
    app = TradutorPremiumApp()
    app.mainloop()
